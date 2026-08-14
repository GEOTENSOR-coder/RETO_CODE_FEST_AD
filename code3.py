import faiss as fb
import pandas as pd
import numpy as np
import json
import pymupdf4llm, pymupdf
from PIL import Image
import magic
import os
import io
import geopandas as gpd
import pytesseract as tesser
from bs4 import BeautifulSoup
import datasets as dats
from transformers import AutoTokenizer
from sentence_transformers import SentenceTransformer
#Para establecer las condiciones de chunking
from langchain_text_splitters import RecursiveCharacterTextSplitter, HTMLHeaderTextSplitter
from langchain_core.documents import Document
import re
from functools import lru_cache
from itertools import islice
import time
import transformers
from openpyxl import load_workbook
import pyogrio
transformers.logging.set_verbosity_error()



def Identifile(file):
    tipefile=magic.from_file(file, mime=True)
    descripe=magic.from_file(file)
    return tipefile, descripe

def Separardf(dataframe, maxtoken=400):
    textoparcial=[]
    tokenstotal=0
    for idx, fila in dataframe.iterrows():
        # Calcula la fila
        parfila=[f"{columna}: {valor}" for columna, valor in fila.items() if pd.notna(valor)]
        textoenfila=" | ".join(parfila) + ".\n"
        tokensdefila= Detextoatoken(textoenfila)
        # Si una fila supera el máximo de tokens por sí sola
        if tokensdefila>maxtoken:
            if textoparcial:
                yield "".join(textoparcial)
                textoparcial=[]
                tokenstotal=0
            #Se divide por los ||
            yield textoenfila
            continue
        # Cerrar el chunk si la nueva fila supera el límite
        if tokenstotal+tokensdefila>maxtoken:
            yield "".join(textoparcial)
            textoparcial=[textoenfila]
            tokenstotal=tokensdefila
        else:
            textoparcial.append(textoenfila)
            tokenstotal+=tokensdefila

    # Si queda algúno restante
    if textoparcial:
        yield "".join(textoparcial)

#El chunkeador cuenta tokens muchas veces sobre las mismas oraciones,
#la cache evita repetir la tokenizacion
@lru_cache(maxsize=200000)
def Detextoatoken(text):
    return len(tokenir.tokenize(text))

#Convierte el documento Json en un diccionario siguiendo las rutas
def Dejsonadiccio(y,out=None,prefijo=''):
    #Si no hay nada
    if out is None:
        out={}
    #En otro caso
    if isinstance(y,dict):
        for k,v in y.items():
            Dejsonadiccio(v,out,f"{prefijo}{k}.")
    elif isinstance(y,list):
        for i,v in enumerate(y):
            Dejsonadiccio(v,out,f"{prefijo}[{i}].")
    else:
        #GUarda el resultado final
        out[prefijo[:-1]]=y
    return out

def Chunckjson(diccio,maxtoken=400):
    #Similar al caso de los csvs
    textoparcial=[]
    tokenstotal=0
    for key,val in diccio.items():
        #Ignora los nulos
        if val is None or str(val).strip()=="":
            continue
        textoenlinea=f"{key}: {val}.\n"
        tokensdelinea=Detextoatoken(textoenlinea)

        #Sí se supera el límite de tokens
        if tokensdelinea>maxtoken:
            if textoparcial:
                yield "".join(textoparcial)
                textoparcial=[]
                tokenstotal=0
            yield textoenlinea
            continue
        #Cerrar el bloque si se supera el límite
        if tokenstotal+tokensdelinea>maxtoken:
            yield "".join(textoparcial)
            textoparcial=[textoenlinea]
            tokenstotal=tokensdelinea
        else:
            textoparcial.append(textoenlinea)
            tokenstotal+=tokensdelinea
    if textoparcial:
        yield "".join(textoparcial)

def leerexcel(doc,chunksize=5000):
    wb=load_workbook(filename=doc,read_only=True,data_only=True)
    sheet=wb.active    
    iterar=sheet.iter_rows(values_only=True)
    header=next(iterar,None)
    #No se detectan los encabezados
    if not header:
        wb.close()
        return
    bloque=[]
    for fila in iterar:
        bloque.append(fila)
        if len(bloque)>=chunksize:
            yield pd.DataFrame(bloque,columns=header)
            bloque=[]
    if bloque:
        yield pd.DataFrame(bloque,columns=header)
    wb.close()


def Readfile(doc,ncarpeta,narchivo,numdoc,fenomeno):
    actuallen=len(Metadatatotal)
    _, ext_os=os.path.splitext(doc.lower())

    #Se inicializa para que las ramas todavia sin terminar (PDF, IMG, PBF)
    #no revienten al asignar campos
    Metadata={}
    #Clasificación de los archivos
    actdoc,descrip=Identifile(doc)
    #Para pdf
    if actdoc=='application/pdf':
        firstid='PDF'
        paraembeding=[]
        numchunk=0
        try:
            #Se abre el documento, se convierte a markdown y se buscan las imágenes
            texto=pymupdf4llm.to_markdown(doc,ocr_language='spa+eng+por')
            #El texto resultante está unificado con toda la inforamción
            haytabla=re.compile(r'(\n(?:\|.*\|\n)+)')
            bloques=haytabla.split(texto)
            for bloque in bloques:
                if not bloque.strip():
                    continue
                #Si es tabla
                if bloque.strip().startswith('|'):
                    chunks=chunkrules.split_datos(bloque)
                else:
                    chunks=chunkrules.split_text(bloque)
                
                for frag in chunks:
                    Metadata={
                        'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                        'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                        'fuente': doc,
                        'formato': firstid,
                        'fenomeno': fenomeno,
                        'posicion': f"{numchunk}",
                        'texto': frag,
                        'fecha': time.ctime(os.path.getctime(doc)),
                        'num_tokens': Detextoatoken(frag)
                    }
                    numchunk+=1
                    paraembeding.append(f"passage: {frag}")
                    Metadatatotal.append(Metadata)
            if paraembeding:
                embeddings=modelot.encode(paraembeding,convert_to_numpy=True, normalize_embeddings=True)
                indicefaiss.add(embeddings)
                
        except Exception as e:
            print(f" ERROR procesando PDF {doc}: {type(e).__name__}: {e}")
            
    elif actdoc in ['image/png', 'image/jpeg', 'image/avif']:
        imagen=Image.open(doc)
        firstid='IMG'
        paraembeding=[]
        numchunk=0
        try:
            imagen=Image.open(doc)
            textototal=tesser.image_to_string(imagen,lang='spa+eng+por')
            textoenimag=textototal.strip()
            if textoenimag:
                chunks=chunkrules.split_text(textoenimag)
                for frag in chunks:
                    Metadata={
                        'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                        'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                        'fuente': doc,
                        'formato': firstid,
                        'fenomeno': fenomeno,
                        'posicion': f"{numchunk}",
                        'texto': frag,
                        'fecha': time.ctime(os.path.getctime(doc)),
                        'num_tokens': Detextoatoken(frag)
                    }
                    paraembeding.append(f"passage: {frag}")
                    Metadatatotal.append(Metadata)
                    numchunk+=1
                if paraembeding:
                    embeddings=modelot.encode(paraembeding,convert_to_numpy=True, normalize_embeddings=True)
                    indicefaiss.add(embeddings)
            else:
                print(f'No se detectó texto en la imagen')
        except Exception as e:
            print(f" ERROR procesando imagen {doc}: {type(e).__name__}: {e}")
    #Para HTML
    elif actdoc=='text/html':
        firstid='HTML'
        paraembeding=[]
        numchunk=0
        
        with open(doc,encoding='utf-8') as f:
            soup=BeautifulSoup(f,'html.parser')            
        #Extra el título
        titulohead=soup.title.string.strip() if soup.title and soup.title.string else "Sin título"
        #Eliminar scripts y estilos
        for dentrode in soup(['script', 'style']):
            dentrode.decompose()
        #Busca imagenes y las reemplaza con el texto obtenido por OCR
        for img in soup.find_all('img'):
            src=img.get('src')
            alt=img.get('alt','Sin descripción')
            textodeimagen=""
            
            if src:
                #Para imagenes locales, guardadas en los documentos
                if not src.startswith(('http://','https://','data:')):
                    ruta=os.path.join(os.path.dirname(doc),src)
                    if os.path.exists(ruta):
                        try:
                            imagen=Image.open(ruta)
                            textodeimagen=tesser.image_to_string(imagen,lang='spa+eng+por').strip()
                        #Si no se puede leer
                        except Exception:
                            pass 
            #Cambia la referencia img por yafue
            yafue=soup.new_tag("p")
            #Reemplaza la imagen por el texto que la describe
            textoreemplazo=" "
            if textodeimagen:
                textoreemplazo+=f". Contenido extraído: {textodeimagen}"
            yafue.string=textoreemplazo
            img.replace_with(yafue)
        #Para las tablas
        for idxt, tabla in enumerate(soup.find_all('table')):
            filasta=[]
            for tr in tabla.find_all('tr'):
                celda=[td.get_text(strip=True) for td in tr.find_all(['td','th'])]
                if celda:
                    filasta.append(" | ".join(celda))
            #Determina el texto en la tabla
            textoentabla=".\n".join(filasta) + "."
            #Crea una etiqueta para reemplazar la tabla por su texto
            nvar=soup.new_tag("p")
            nvar.string=f"{textoentabla}"
            tabla.replace_with(nvar)
        #Se hace chunking del documento completo convertido en texto
        htmldata=str(soup.body) if soup.body else str(soup)
        htmldivid=HTMLHeaderTextSplitter(headers_to_split_on=[("h1","H1"), ("h2","H2"),("h3","H3"),("h4","H4")])
        htmlsplit=htmldivid.split_text(htmldata)
        chunks=chunkrules.split_documents(htmlsplit)
        
        for frag in chunks:
            contenido=frag.page_content
            #Si hay subtitutlos 
            cheaders=" > ".join([f"{v}" for k,v in frag.metadata.items()])
            if cheaders:
                textototal=f"Documento: {titulohead}. Sección: {cheaders}. Contenido: {contenido}"
            else:
                textototal=f"Documento: {titulohead}. Contenido: {contenido}"
                
            Metadata={
                'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                'fuente': doc, 'formato': firstid, 'fenomeno': fenomeno,
                'posicion': f"{numchunk}",
                'texto': textototal,
                'fecha': time.ctime(os.path.getctime(doc)),
                'num_tokens': Detextoatoken(textototal)            }
            paraembeding.append(f"passage: {textototal}")
            Metadatatotal.append(Metadata)
            numchunk+= 1
            
        if paraembeding:
            embeddings=modelot.encode(paraembeding,convert_to_numpy=True, normalize_embeddings=True)
            indicefaiss.add(embeddings)
            
    # Para CSV
    elif actdoc=='text/csv':
        paraembeding=[]
        firstid='CSV'
        numchunk=0
        maximodelineas=1000
        for dfsdiv in pd.read_csv(doc, chunksize=1000):
            df=dfsdiv.dropna(axis=1, how='all')
            for textofinal in (Separardf(df)):
               chunks=chunkrules.split_datos(textofinal)
               for frag in (chunks):
                   Metadata={'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                       'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                       'fuente': doc,
                       'formato': firstid,
                       'fenomeno': fenomeno,
                       'posicion': f"{numchunk}",
                       'texto': frag,
                       'fecha': time.ctime(os.path.getctime(doc)),
                       'num_tokens': Detextoatoken(frag)
                   }
            
                   paraembeding.append(f"passage: {frag}")
                   Metadatatotal.append(Metadata)
                   numchunk+=1
                   #Seguridad para no sobrecargar la memoria
                   if len(paraembeding)>=maximodelineas:
                        embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                        indicefaiss.add(embeddings)
                        paraembeding=[]
            if paraembeding:
                embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                indicefaiss.add(embeddings)
    #Para excel
    #Si lo detecta automaticamente
    elif actdoc=='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':
        paraembeding=[]
        firstid='XLSX'
        numchunk=0
        maximodelineas=1000
        for dfsdiv in leerexcel(doc,chunksize=1000):
            df=dfsdiv.dropna(axis=1, how='all')
            for textofinal in (Separardf(df)):
                  chunks=chunkrules.split_datos(textofinal)
                  for frag in (chunks):
                      Metadata={'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                          'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                          'fuente': doc,
                          'formato': firstid,
                          'fenomeno': fenomeno,
                          'posicion': f"{numchunk}",
                          'texto': frag,
                          'fecha': time.ctime(os.path.getctime(doc)),
                          'num_tokens': Detextoatoken(frag)
                      }
                      numchunk+=1
                      paraembeding.append(f"passage: {frag}")
                      Metadatatotal.append(Metadata)
                      #Seguridad para no sobrecargar la memoria
                      if len(paraembeding)>=maximodelineas:
                        embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                        indicefaiss.add(embeddings)
                        paraembeding=[]
        if paraembeding:
            embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
            indicefaiss.add(embeddings)
    #Si lo detecta como un zip
    elif actdoc=='application/zip' and doc.lower().endswith('.xlsx'):
        paraembeding=[]
        firstid='XLSX'
        numchunk=0
        maximodelineas=1000
        for dfsdiv in leerexcel(doc,chunksize=1000):
            df=dfsdiv.dropna(axis=1, how='all')
            for textofinal in (Separardf(df)):
                  chunks=chunkrules.split_datos(textofinal)
                  for frag in (chunks):
                      Metadata={'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                          'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                          'fuente': doc,
                          'formato': firstid,
                          'fenomeno': fenomeno,
                          'posicion': f"{numchunk}",
                          'texto': frag,
                          'fecha': time.ctime(os.path.getctime(doc)),
                          'num_tokens': Detextoatoken(frag)
                      }
                      numchunk+=1
                      paraembeding.append(f"passage: {frag}")
                      Metadatatotal.append(Metadata)
                      #Seguridad para no sobrecargar la memoria
                      if len(paraembeding)>=maximodelineas:
                        embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                        indicefaiss.add(embeddings)
                        paraembeding=[]
        if paraembeding:
            embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
            indicefaiss.add(embeddings)
    #Para Json
    elif actdoc=='application/json':
        paraembeding=[]
        firstid='JSON'
        numchunk=0
        #Metadata['formato']=firstid
        with open(doc) as f:
            datos=json.load(f)
        jsonresum=Dejsonadiccio(datos)
        for textofinal in (Chunckjson(jsonresum)):
            #Divide los chunks del texto
            chunks=chunkrules.split_datos(textofinal)
            for frag in (chunks):
                Metadata = {
                    'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                    'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                    'fuente': doc,
                    'formato': firstid,
                    'fenomeno': fenomeno,
                    'posicion': f"{numchunk}",
                    'texto': frag,
                    'fecha': time.ctime(os.path.getctime(doc)),
                    'num_tokens': Detextoatoken(frag)
                }
                numchunk+=1
                paraembeding.append(f"passage: {frag}")
                Metadatatotal.append(Metadata)

        if paraembeding:
            embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
            indicefaiss.add(embeddings)

    #Para txt y markdown
    elif actdoc=='text/plain':
        #Para asegurar que sea markdown
        _, osactdoc=os.path.splitext(doc)
        if osactdoc in ['.md', '.markdown', '.mdown', '.mkdn', '.mkd', '.mdwn']:
            paraembeding=[]
            firstid='MRKD'
            numchunk=0
            with open(doc,'r',encoding='utf-8') as f:
                content=f.read()
            #Busca tablas dentro del documento
            haytabla=re.compile(r'(\n(?:\|.*\|\n)+)')
            bloques=haytabla.split(content)
            #Verifica que hayan datos
            for bloque in bloques:
                if not bloque.strip():
                    continue
                #Si es tabla
                if bloque.strip().startswith('|'):
                    chunks=chunkrules.split_datos(bloque)
                else:
                    chunks=chunkrules.split_text(bloque)
                for frag in (chunks):
                    Metadata={
                        'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                        'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                        'fuente': doc,
                        'formato': firstid,
                        'fenomeno': fenomeno,
                        'posicion': f"{numchunk}",
                        'texto': frag,
                        'fecha': time.ctime(os.path.getctime(doc)),
                        'num_tokens': Detextoatoken(frag)
                    }
                    numchunk+=1
                    paraembeding.append(f"passage: {frag}")
                    Metadatatotal.append(Metadata)
            if paraembeding:
                embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                indicefaiss.add(embeddings)
        elif osactdoc=='.txt':
            paraembeding=[]
            firstid='TXT'
            numchunk=0
            with open(doc,'r',encoding='utf-8') as f:
                content=f.read()
            #Busca tablas dentro del documento
            haytabla=re.compile(r'(\n(?:\|.*\|\n)+)')
            bloques=haytabla.split(content)
            #Verifica que hayan datos
            for bloque in bloques:
                if not bloque.strip():
                    continue
                #Si es tabla
                if bloque.strip().startswith('|'):
                    chunks=chunkrules.split_datos(bloque)
                else:
                    chunks=chunkrules.split_text(bloque)
                for frag in (chunks):
                    Metadata={
                        'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                        'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                        'fuente': doc,
                        'formato': firstid,
                        'fenomeno': fenomeno,
                        'posicion': f"{numchunk}",
                        'texto': frag,
                        'fecha': time.ctime(os.path.getctime(doc)),
                        'num_tokens': Detextoatoken(frag)
                    }
                    numchunk+=1
                    paraembeding.append(f"passage: {frag}")
                    Metadatatotal.append(Metadata)

            if paraembeding:
                embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                indicefaiss.add(embeddings)

        #Si un csv es texto plano
        elif osactdoc=='.csv':
            paraembeding=[]
            firstid='CSV'
            numchunk=0
            maximodelineas=1000
            for dfsdiv in pd.read_csv(doc, chunksize=1000):
                df=dfsdiv.dropna(axis=1, how='all')
                for textofinal in (Separardf(df)):
                   chunks=chunkrules.split_datos(textofinal)
                   for frag in (chunks):
                       Metadata={'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                           'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                           'fuente': doc,
                           'formato': firstid,
                           'fenomeno': fenomeno,
                           'posicion': f"{numchunk}",
                           'texto': frag,
                           'fecha': time.ctime(os.path.getctime(doc)),
                           'num_tokens': Detextoatoken(frag)
                       }
                
                       paraembeding.append(f"passage: {frag}")
                       Metadatatotal.append(Metadata)
                       numchunk+=1
                       #Seguridad para no sobrecargar la memoria
                       if len(paraembeding)>=maximodelineas:
                            embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                            indicefaiss.add(embeddings)
                            paraembeding=[]
                if paraembeding:
                    embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                    indicefaiss.add(embeddings)
    #Para PBF
    elif actdoc=='application/octet-stream' or 'OpenStreetMap Protocolbuffer' in descrip or ext_os == '.pbf':
        #Verificando que si es pbf
        nombredoc,osactdoc=os.path.splitext(doc.lower())
        _, osactdoc2=os.path.splitext(nombredoc)
        if (osactdoc2+osactdoc)=='.osm.pbf' or osactdoc=='.pbf':
            firstid='PBF'
            paraembeding=[]
            numchunk=0 
            works=False          
            #Determinar cuales son las capas
            try:
                incapas=pyogrio.list_layers(doc)
                capas=[capa[0] for capa in incapas]
            except Exception as e:
                print(f" ERROR al leer el listado de capas del PBF: {type(e).__name__}: {e}")
                capas=[]
            #Sobre cada capa
            for kl in capas:
                try:
                    gdf=gpd.read_file(doc, engine="pyogrio", layer=kl)
                    works=True
                    
                    if gdf.empty:
                        continue
                    if 'geometry' in gdf.columns:
                        gdf=gdf.drop(columns=['geometry'])
                    gdf=gdf.dropna(axis=1,how='all')
                    
                    for textofinal in Separardf(gdf):
                        chunks=chunkrules.split_datos(textofinal)
                        for frag in chunks:
                            textocompleto=f"Capa [{kl}]: {frag}"
                            
                            Metadata={
                                'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                                'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                                'fuente': doc,
                                'formato': firstid,
                                'fenomeno': fenomeno,
                                'posicion': f"{numchunk}",
                                'texto': textocompleto,
                                'fecha': time.ctime(os.path.getctime(doc)),
                                'num_tokens': Detextoatoken(textocompleto)
                            }
                            numchunk+=1
                            paraembeding.append(f"passage: {textocompleto}")
                            Metadatatotal.append(Metadata)
                except Exception as e:
                    pass 

            if paraembeding:
                embeddings=modelot.encode(paraembeding,convert_to_numpy=True, normalize_embeddings=True)
                indicefaiss.add(embeddings)
                
            #Leerlo como tabla en el peor de los casos
            if not works:
                try:
                    maximodelineas=1000
                    for dfsdiv in pd.read_csv(doc, chunksize=1000, sep=None, engine='python', on_bad_lines='skip', encoding='latin-1', quoting=3):
                        df=dfsdiv.dropna(axis=1, how='all')
                        for textofinal in (Separardf(df)):
                            chunks=chunkrules.split_datos(textofinal)
                            for frag in (chunks):
                                Metadata={
                                    'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                                    'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                                    'fuente': doc,
                                    'formato': f"{firstid}-TXT",
                                    'fenomeno': fenomeno,
                                    'posicion': f"{numchunk}",
                                    'texto': frag,
                                    'fecha': time.ctime(os.path.getctime(doc)),
                                    'num_tokens': Detextoatoken(frag)
                                }
                                numchunk+=1
                                paraembeding.append(f"passage: {frag}")
                                Metadatatotal.append(Metadata)
                                if len(paraembeding)>=maximodelineas:
                                    embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                                    indicefaiss.add(embeddings)
                                    paraembeding=[]
                except Exception as e:
                    print(f" ERROR: El PBF no se pudo procesar como mapa ni como tabla: {type(e).__name__}: {e}")
            #Se intenta buscar información en el archivo, considerando que está en español, inglés o portugués 
            if numchunk==0:
                try:
                    #Se abre como si fuera un Markdown
                    with open(doc, 'r', encoding='utf-8', errors='ignore') as f:
                        intento=f.read(1000).lower()
                    #Palabras para determinar el idioma
                    palabras={' el ', ' la ', ' de ', ' que ', ' y ', ' a ', ' en ', ' un ', ' con ', # Español
                            ' the ', ' be ', ' to ', ' of ', ' and ', ' in ', ' that ', ' for ',     # Inglés
                            ' o ', ' a ', ' de ', ' e ', ' do ', ' da ', ' que ', ' em ', ' um '}     # Portugués
                    #SI las palabras están en el documento se guarda su valor
                    coincidencias=sum(1 for palabra in palabras if palabra in intento)
                    #SI hay más de 3 palabras que coincidan
                    if coincidencias >= 3:
                        firstid='pbf'
                        paraembeding=[]
                        numchunk=0
                        #Se lee el documento como markdown
                        with open(doc, 'r', encoding='utf-8', errors='ignore') as f:
                            content=f.read()
                        haytabla=re.compile(r'(\n(?:\|.*\|\n)+)')
                        bloques=haytabla.split(content)
                        for bloque in bloques:
                            if not bloque.strip():
                                continue
                            if bloque.strip().startswith('|'):
                                chunks=chunkrules.split_datos(bloque)
                            else:
                                chunks=chunkrules.split_text(bloque)
                                    
                            for frag in chunks:
                                Metadata={
                                        'doc_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}",
                                        'chunk_id': f"{ncarpeta}-{narchivo}-{firstid}-{numdoc}-Chunk-{numchunk}",
                                        'fuente': doc,
                                        'formato': firstid,
                                        'fenomeno': fenomeno,
                                        'posicion': f"{numchunk}",
                                        'texto': frag,
                                        'fecha': time.ctime(os.path.getctime(doc)),
                                        'num_tokens': Detextoatoken(frag)                        }
                                numchunk+=1
                                paraembeding.append(f"passage: {frag}")
                                Metadatatotal.append(Metadata)
                                    
                        if paraembeding:
                            embeddings=modelot.encode(paraembeding, convert_to_numpy=True, normalize_embeddings=True)
                            indicefaiss.add(embeddings)
                        print(f"  [!] Archivo fragmentado por coincidencias de idioma")
                    else:
                        print(f"  [X] Archivo omitido por contenido ilegible:")

                        if paraembeding:
                            embeddings=modelot.encode(paraembeding,convert_to_numpy=True, normalize_embeddings=True)
                            indicefaiss.add(embeddings)
                except Exception as e:
                    print(f" ERROR al intentar leer el archivo {doc}: {type(e).__name__}: {e}")
                
        else:
            print('Formato de pbf incorrecto.')

    else:
        print('Formato no soportado')
    print(f"   -> Total fragmentos: {len(Metadatatotal)}, fragmentos de este archivo: {len(Metadatatotal)-actuallen}")


#------------------------------------------------------------------
# CHUNKING ORACIONAL
# El reto exige que ningun fragmento tenga oraciones incompletas:
# los cortes solo pueden caer en un limite de oracion. Por eso ya no
# se corta por caracteres, sino que se arma el chunk juntando
# oraciones completas hasta llegar al limite de tokens.
#------------------------------------------------------------------

#Palabras que llevan punto pero NO terminan la oracion
ABREVIATURAS={
    'sr','sra','srta','dr','dra','ing','lic','msc','mg','ph.d','phd','vs','etc','ej',
    'p.ej','fig','tab','num','núm','art','pag','pág','pp','vol','ed','av','aprox',
    'max','máx','min','mín','mr','mrs','ms','st','jr','inc','ltd','co','al','i.e',
    'e.g','cf','approx','dept','univ','a.m','p.m','ee.uu','ee','uu','u.s','u.k',
    'no','nro','fig','ref','op','cit','pe','sgto','gral','cnel','tte','cap'
}
#Fin de oracion: signo terminal + comillas/parentesis de cierre + espacio o salto
FINORACION=re.compile(r'(?<=[.!?…])["»”’\')\]]*\s+')
#Ultima palabra antes del signo terminal (para descartar abreviaturas)
ULTPALABRA=re.compile(r'([^\s]+?)[.!?…]$')

def Contarpalabras(texto):
    return len(texto.split())

def Enoraciones(texto):
    #Devuelve la lista de oraciones COMPLETAS de un texto
    texto=texto.strip()
    if not texto:
        return []
    oraciones=[]
    inicio=0
    for m in FINORACION.finditer(texto):
        previo=texto[:m.start()]
        ult=ULTPALABRA.search(previo)
        if ult:
            palabra=ult.group(1).lower()
            #Abreviaturas, iniciales sueltas y numeraciones no cierran oracion
            if palabra in ABREVIATURAS or palabra.rstrip('.') in ABREVIATURAS:
                continue
            if len(palabra)==1 and palabra.isalpha():
                continue
            #Un numero solo es enumeracion ("1. Item") si abre la linea;
            #si cierra la frase (un año, una cifra) si termina la oracion
            if palabra.replace(',','').isdigit():
                antes=previo[:ult.start()]
                if antes=='' or antes.rstrip(' \t').endswith('\n'):
                    continue
        oracion=texto[inicio:m.end()].strip()
        if oracion:
            oraciones.append(oracion)
        inicio=m.end()
    resto=texto[inicio:].strip()
    if resto:
        oraciones.append(resto)
    return oraciones

class Chunkeador:
    #max_tokens: limite del encoder (512), se deja margen para el prefijo "passage: "
    #max_palabras: limite de 250 palabras que exige el formato de salida
    #solape: cuantas ORACIONES completas se repiten entre chunks vecinos
    def __init__(self,max_tokens=400,max_palabras=240,solape=1,min_tokens=15):
        self.max_tokens=max_tokens
        self.max_palabras=max_palabras
        self.solape=solape
        self.min_tokens=min_tokens
        #Solo se usa cuando UNA sola "oracion" ya pasa el limite
        #(filas de tablas, listas sin puntuacion, texto de OCR sin puntos)
        self.emergencia=RecursiveCharacterTextSplitter(
            chunk_size=max_tokens,
            chunk_overlap=0,
            length_function=Detextoatoken,
            keep_separator='end',  #el separador queda al FINAL del trozo, no al inicio
            separators=['\n',' | ','; ',': ',', ',' ','']
        )

    def _porpalabras(self,texto):
        palabras=texto.split()
        return [" ".join(palabras[i:i+self.max_palabras])
                for i in range(0,len(palabras),self.max_palabras)]

    def _unidades(self,texto):
        #Bloques por linea en blanco -> y dentro de cada bloque, oraciones completas
        unidades=[]
        for bloque in re.split(r'\n\s*\n',texto):
            if not bloque.strip():
                continue
            for oracion in Enoraciones(bloque):
                if Detextoatoken(oracion)<=self.max_tokens and Contarpalabras(oracion)<=self.max_palabras:
                    unidades.append(oracion)
                    continue
                #Caso extremo: no hay limite oracional donde cortar
                for trozo in self.emergencia.split_text(oracion):
                    if not trozo.strip():
                        continue
                    if Contarpalabras(trozo)>self.max_palabras:
                        unidades.extend(self._porpalabras(trozo))
                    else:
                        unidades.append(trozo.strip())
        return unidades

    def split_text(self,texto):
        if not texto or not str(texto).strip():
            return []
        unidades=self._unidades(str(texto))
        chunks=[]
        actual=[]
        ntok=0
        npal=0
        for u in unidades:
            tu=Detextoatoken(u)
            pu=Contarpalabras(u)
            if actual and (ntok+tu>self.max_tokens or npal+pu>self.max_palabras):
                chunks.append(" ".join(actual))
                #El solape se hace con oraciones enteras, nunca a mitad de frase
                actual=actual[-self.solape:] if self.solape>0 else []
                #Si el solape no deja espacio para la oracion nueva, se recorta
                while actual and (sum(Detextoatoken(x) for x in actual)+tu>self.max_tokens
                                  or sum(Contarpalabras(x) for x in actual)+pu>self.max_palabras):
                    actual.pop(0)
                ntok=sum(Detextoatoken(x) for x in actual)
                npal=sum(Contarpalabras(x) for x in actual)
            actual.append(u)
            ntok+=tu
            npal+=pu
        if actual:
            chunks.append(" ".join(actual))
        #Los fragmentos muy cortos (titulos sueltos, colas de texto) se pegan
        #al vecino siempre que no se pasen de los limites
        fusionados=[]
        for c in chunks:
            if fusionados and (Detextoatoken(c)<self.min_tokens
                               or Detextoatoken(fusionados[-1])<self.min_tokens):
                fusion=fusionados[-1]+" "+c
                if Detextoatoken(fusion)<=self.max_tokens and Contarpalabras(fusion)<=self.max_palabras:
                    fusionados[-1]=fusion
                    continue
            fusionados.append(c)
        return [c.strip() for c in fusionados if c.strip()]
    def split_datos(self, texto):
        #Ignora la gramática al tratarse de tablas
        if not texto or not str(texto).strip():
            return []
        #Se revisa qeu los bloques no pasen el número de tokens ni de palabras
        if Detextoatoken(texto)<=self.max_tokens and Contarpalabras(texto)<=self.max_palabras:
            return [texto.strip()]
        #Si la fila es muy grande se corta
        fragmentos=self.emergencia.split_text(texto)
        return [c.strip() for c in fragmentos if c.strip()]
    #Misma firma que RecursiveCharacterTextSplitter para el caso del HTML
    def split_documents(self,documentos):
        salida=[]
        for d in documentos:
            for trozo in self.split_text(d.page_content):
                salida.append(Document(page_content=trozo,metadata=dict(d.metadata)))
        return salida

#Inicio del transformer
transformer='intfloat/multilingual-e5-small'
tokenir=AutoTokenizer.from_pretrained(transformer)
#Reglas para el chunking (se crea despues del tokenizador porque lo necesita)
chunkrules=Chunkeador(max_tokens=400,max_palabras=240,solape=1)
modelot=SentenceTransformer(transformer)
#Dimensión del vector que produce el modelo
lonvector=384
indicefaiss=fb.IndexFlatIP(lonvector)
carpetacontenedor='.'
#Guarda la relación entre los datos del indice Faiss
#y el orden de carga de los archivos
Metadatatotal=[]


EXCLUIDAS={'venv','.venv','env','.git','__pycache__'}
EXTIGNORAR={'.py','.pyc','.faiss','.jsonl','.log','.DS_Store'}

numdoc=0
for carpeta, subcarpetas, archivos in os.walk(carpetacontenedor):
    subcarpetas[:]=[s for s in subcarpetas if s not in EXCLUIDAS]
    if not archivos:
        continue
    rutarel=os.path.relpath(carpeta,carpetacontenedor).upper()
    if 'F1_' in rutarel:
        fenomeno=1
    elif 'F2_' in rutarel:
        fenomeno=2
    elif 'F3_' in rutarel:
        fenomeno=3
    else:
        fenomeno=0
    nombrecarpeta=os.path.basename(carpeta)
    if nombrecarpeta=='.' or nombrecarpeta=='':
        nombrecarpeta='BASE'
    idcarpeta=nombrecarpeta[:4].upper()
    for archivo in archivos:
        if os.path.splitext(archivo)[1].lower() in EXTIGNORAR:
            continue
        numdoc+=1
        idarch=os.path.basename(archivo)[:4].upper()
        rutadelarchivo=os.path.join(carpeta,archivo)
        print(f"[{numdoc}] {rutadelarchivo}")
        try:
            Readfile(rutadelarchivo,idcarpeta,idarch,numdoc,fenomeno)
        except Exception as e:
            print(f" ERROR en {rutadelarchivo}: {type(e).__name__}: {e}")

with open("Metada.jsonl", 'w', encoding='utf-8') as file:
    for item in Metadatatotal:
        #Convierte el diccionario a string
        linea=json.dumps(item,ensure_ascii=False)
        file.write(linea+'\n')
#Se guarda el indice faiss
fb.write_index(indicefaiss,'indice.faiss')

